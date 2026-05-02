# -*- coding: utf-8 -*-
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook(
    r"D:\Work\Driver Survey\03) z. All g 52 - Routine.xlsx",
    read_only=True, data_only=False
)

# ── Survey sheet headers ─────────────────────────────────────────────────────
ws = wb["Survey"]
headers = {}
for cell in next(ws.iter_rows(min_row=1, max_row=1)):
    if cell.value:
        headers[cell.column] = cell.value

print("=== KEY SURVEY COLUMNS ===")
for col_letter in ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P",
                   "BK","BL","BM","BN","BO","BP","BQ","BR","BS","BT","BU","BV","BW"]:
    idx = column_index_from_string(col_letter)
    print("  %s (%d): %s" % (col_letter, idx, headers.get(idx, "(empty)")))

print()
print("=== ALL SURVEY HEADERS ===")
for idx in sorted(headers):
    print("  %s (%d): %s" % (get_column_letter(idx), idx, headers[idx]))

# ── #5,#6 sheet: row 1 values and formulas ──────────────────────────────────
print()
print("=== #5,#6 SHEET ROW 1 (reference values) ===")
ws56 = wb["#5,#6"]
for row in ws56.iter_rows(min_row=1, max_row=3, max_col=60):
    vals = [(get_column_letter(c.column), c.value) for c in row if c.value is not None]
    for col, v in vals:
        print("  [%s]: %s" % (col, v))
    print()

print()
print("=== #5,#6 FORMULAS (row 8 = first data row, cols E onwards) ===")
for row in ws56.iter_rows(min_row=8, max_row=8, max_col=60):
    for c in row:
        if c.value is not None and str(c.value).startswith("="):
            print("  [%s]: %s" % (get_column_letter(c.column), c.value))
